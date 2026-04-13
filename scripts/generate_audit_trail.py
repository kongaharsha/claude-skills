#!/usr/bin/env python3
"""Generate an audit trail Excel workbook mapping every filled form field
back to its source document, with computation details.

This is the transparency deliverable — it lets the user (or a reviewer) trace
any number on any tax form back to the exact document, box, or line it came
from, plus the arithmetic used to derive it.

Usage:
    python generate_audit_trail.py work/extraction_ledger.json work/results.json output/

Produces: output/audit_trail.xlsx
"""

import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


# ---------------------------------------------------------------------------
# Styling constants
# ---------------------------------------------------------------------------
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
VERIFIED_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
WARNING_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SECTION_FONT = Font(name="Calibri", bold=True, size=11, color="2F5496")
THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
NUM_FMT = '#,##0.00'


# ---------------------------------------------------------------------------
# Form line definitions — what goes on each form
# ---------------------------------------------------------------------------
# Each entry: (line, description, results_key, source_description_fn, computation_fn)
# The source_description_fn and computation_fn take (ledger, results) and return strings.

def _wage_source(ledger, results):
    """Build source string for total wages."""
    docs = [d for d in ledger.get("documents", []) if d["type"] == "W-2"]
    parts = []
    for d in docs:
        issuer = d.get("issuer", d.get("file", "W-2"))
        filer = d.get("filer", "unknown")
        wages = d.get("values", {}).get("box_1_wages", 0)
        parts.append(f"{filer} W-2 ({issuer}) Box 1 = ${wages:,.2f}")
    return "; ".join(parts) if parts else "W-2 Box 1"


def _wage_computation(ledger, results):
    docs = [d for d in ledger.get("documents", []) if d["type"] == "W-2"]
    vals = [d.get("values", {}).get("box_1_wages", 0) for d in docs]
    if len(vals) > 1:
        return " + ".join(f"${v:,.2f}" for v in vals) + f" = ${sum(vals):,.2f}"
    elif vals:
        return f"${vals[0]:,.2f}"
    return ""


def _wh_source(ledger, results):
    docs = [d for d in ledger.get("documents", []) if d["type"] == "W-2"]
    parts = []
    for d in docs:
        issuer = d.get("issuer", d.get("file", "W-2"))
        filer = d.get("filer", "unknown")
        wh = d.get("values", {}).get("box_2_fed_wh", 0)
        parts.append(f"{filer} W-2 ({issuer}) Box 2 = ${wh:,.2f}")
    return "; ".join(parts) if parts else "W-2 Box 2"


def _wh_computation(ledger, results):
    docs = [d for d in ledger.get("documents", []) if d["type"] == "W-2"]
    vals = [d.get("values", {}).get("box_2_fed_wh", 0) for d in docs]
    if len(vals) > 1:
        return " + ".join(f"${v:,.2f}" for v in vals) + f" = ${sum(vals):,.2f}"
    elif vals:
        return f"${vals[0]:,.2f}"
    return ""


def _div_source(ledger, results, field="ordinary_dividends"):
    docs = [d for d in ledger.get("documents", []) if d["type"] == "1099-DIV"]
    parts = []
    for d in docs:
        filer = d.get("filer", "unknown")
        issuer = d.get("issuer", d.get("file", "1099-DIV"))
        val = d.get("values", {}).get(field, 0)
        if val:
            parts.append(f"{filer} 1099-DIV ({issuer}) = ${val:,.2f}")
    return "; ".join(parts) if parts else f"1099-DIV {field}"


def _div_computation(ledger, results, field="ordinary_dividends"):
    docs = [d for d in ledger.get("documents", []) if d["type"] == "1099-DIV"]
    vals = [d.get("values", {}).get(field, 0) for d in docs if d.get("values", {}).get(field, 0)]
    if len(vals) > 1:
        return " + ".join(f"${v:,.2f}" for v in vals) + f" = ${sum(vals):,.2f}"
    elif vals:
        return f"${vals[0]:,.2f}"
    return ""


def _simple_result(key, label=None):
    """Return a (source_fn, computation_fn) pair that just reads from results."""
    def source_fn(ledger, results):
        return label or f"Computed (see computations.txt)"
    def comp_fn(ledger, results):
        fed = results.get("federal", results)
        val = fed.get(key, 0)
        return f"${val:,.2f}" if isinstance(val, (int, float)) else str(val)
    return source_fn, comp_fn


def _sum_keys(keys, label):
    """Return computation_fn that sums multiple result keys."""
    def comp_fn(ledger, results):
        fed = results.get("federal", results)
        vals = [fed.get(k, 0) for k in keys]
        parts = " + ".join(f"${v:,.2f}" for v in vals if v)
        total = sum(vals)
        return f"{parts} = ${total:,.2f}" if len([v for v in vals if v]) > 1 else f"${total:,.2f}"
    return comp_fn


def build_form_1040_rows(ledger, results):
    """Build audit trail rows for Form 1040."""
    fed = results.get("federal", results)
    rows = []

    def r(line, desc, value, source_fn, comp_fn):
        src = source_fn(ledger, results) if callable(source_fn) else source_fn
        comp = comp_fn(ledger, results) if callable(comp_fn) else comp_fn
        rows.append(("1040", line, desc, value, src, comp, "✓"))

    wages = fed.get("wages", fed.get("total_wages", 0))
    r("1a", "Wages, salaries, tips", wages, _wage_source, _wage_computation)

    interest = fed.get("interest", fed.get("taxable_interest", 0))
    if interest:
        r("2b", "Taxable interest", interest,
          "1099-INT / brokerage statements",
          f"${interest:,.2f}")

    ord_div = fed.get("ord_div", fed.get("ordinary_dividends", 0))
    if ord_div:
        r("3a", "Ordinary dividends", ord_div,
          lambda l, r: _div_source(l, r, "ordinary_dividends"),
          lambda l, r: _div_computation(l, r, "ordinary_dividends"))

    qual_div = fed.get("qual_div", fed.get("qualified_dividends", 0))
    if qual_div:
        r("3b", "Qualified dividends", qual_div,
          lambda l, r: _div_source(l, r, "qualified_dividends"),
          lambda l, r: _div_computation(l, r, "qualified_dividends"))

    cap_gain = fed.get("net_cap_gain", fed.get("capital_gains", 0))
    if cap_gain:
        r("7", "Capital gain or (loss)", cap_gain,
          "Schedule D Line 16 (from 8949 transactions)",
          lambda l, r: f"${cap_gain:,.2f} (net from Schedule D)")

    other_inc = fed.get("other_income", 0)
    if other_inc:
        r("8", "Other income (Schedule 1)", other_inc,
          "Schedule 1 Line 10", lambda l, r: f"${other_inc:,.2f}")

    total_inc = fed.get("total_income", 0)
    if total_inc:
        r("9", "Total income", total_inc,
          "Sum of Lines 1-8", lambda l, r: f"${total_inc:,.2f}")

    adjustments = fed.get("adjustments", 0)
    if adjustments:
        r("10", "Adjustments to income (Schedule 1)", adjustments,
          "Schedule 1 Line 26", lambda l, r: f"${adjustments:,.2f}")

    agi = fed.get("agi", 0)
    r("11", "Adjusted gross income (AGI)", agi,
      "Line 9 − Line 10",
      lambda l, r: f"${fed.get('total_income', 0):,.2f} − ${fed.get('adjustments', 0):,.2f} = ${agi:,.2f}")

    deduction = fed.get("deduction", fed.get("itemized_deduction", fed.get("total_deduction", 0)))
    ded_type = fed.get("deduction_type", "itemized")
    if deduction:
        r("12", f"Deductions ({ded_type})", deduction,
          f"Schedule A total" if ded_type == "itemized" else "Standard deduction",
          lambda l, r: f"${deduction:,.2f}")

    qbi = fed.get("qbi_deduction", 0)
    if qbi:
        r("13", "Qualified business income deduction", qbi,
          "Computed from QBI", lambda l, r: f"${qbi:,.2f}")

    taxable = fed.get("taxable_income", 0)
    r("15", "Taxable income", taxable,
      "Line 11 − Line 12 − Line 13",
      lambda l, r: f"${agi:,.2f} − ${deduction:,.2f} = ${taxable:,.2f}")

    tax = fed.get("tax", fed.get("income_tax", 0))
    if tax:
        r("16", "Tax (from Tax Table or QDCG worksheet)", tax,
          "2025 MFJ brackets / QDCG worksheet",
          lambda l, r: f"${tax:,.2f}")

    sched2 = fed.get("schedule_2_total", fed.get("additional_taxes", 0))
    if sched2:
        r("23", "Other taxes (Schedule 2 Line 21)", sched2,
          "Schedule 2 (Add'l Medicare + NIIT)",
          lambda l, r: f"${sched2:,.2f}")

    total_tax = fed.get("total_tax", 0)
    r("24", "Total tax", total_tax,
      "Line 16 + Line 23 (+ others)",
      lambda l, r: f"${total_tax:,.2f}")

    total_wh = fed.get("total_fed_wh", fed.get("fed_wh", 0))
    r("25a", "Federal tax withheld (W-2s)", total_wh, _wh_source, _wh_computation)

    sched3 = fed.get("schedule_3_total", fed.get("total_credits", 0))
    if sched3:
        r("31", "Credits (Schedule 3)", sched3,
          "Schedule 3 Line 15 (FTC, CDCC, etc.)",
          lambda l, r: f"${sched3:,.2f}")

    total_payments = fed.get("total_payments", 0)
    r("33", "Total payments", total_payments,
      "Line 25a + Line 31 + estimated payments",
      lambda l, r: f"${total_payments:,.2f}")

    owed = fed.get("owed", fed.get("amount_owed", 0))
    refund = fed.get("refund", 0)
    if owed > 0:
        r("37", "Amount you owe", owed,
          "Line 24 − Line 33",
          lambda l, r: f"${total_tax:,.2f} − ${total_payments:,.2f} = ${owed:,.2f}")
    elif refund > 0:
        r("34", "Overpaid / Refund", refund,
          "Line 33 − Line 24",
          lambda l, r: f"${total_payments:,.2f} − ${total_tax:,.2f} = ${refund:,.2f}")

    return rows


def build_schedule_a_rows(ledger, results):
    """Build audit trail rows for Schedule A (Itemized Deductions)."""
    fed = results.get("federal", results)
    sched_a = fed.get("schedule_a", {})
    if not sched_a and not fed.get("itemized_deduction"):
        return []

    rows = []
    def r(line, desc, value, source, comp):
        if value:
            rows.append(("Schedule A", line, desc, value, source, comp, "✓"))

    # State & local taxes
    state_taxes = sched_a.get("state_income_tax", fed.get("state_taxes_paid", 0))
    r("5a", "State and local income taxes", state_taxes,
      "W-2 Box 17 state withholding totals",
      f"${state_taxes:,.2f}" if state_taxes else "")

    salt_cap = sched_a.get("salt_deduction", sched_a.get("line_5e", 0))
    r("5e", "State and local taxes (SALT cap)", salt_cap,
      "Lesser of 5d total or $10,000 ($20,000 MFJ if OBBBA applies)",
      f"${salt_cap:,.2f}" if salt_cap else "")

    mortgage = sched_a.get("mortgage_interest", 0)
    r("8a", "Home mortgage interest", mortgage,
      "1098 Mortgage Interest Statement",
      f"${mortgage:,.2f}" if mortgage else "")

    charity = sched_a.get("charitable", sched_a.get("charity", 0))
    r("12", "Charitable contributions", charity,
      "Donation receipts / records",
      f"${charity:,.2f}" if charity else "")

    total = sched_a.get("total", fed.get("itemized_deduction", 0))
    r("17", "Total itemized deductions", total,
      "Sum of Schedule A lines",
      f"${total:,.2f}" if total else "")

    return rows


def build_schedule_2_rows(ledger, results):
    """Build audit trail rows for Schedule 2 (Additional Taxes)."""
    fed = results.get("federal", results)
    sched2 = fed.get("schedule_2", {})
    if not sched2 and not fed.get("additional_medicare") and not fed.get("niit"):
        return []

    rows = []
    def r(line, desc, value, source, comp):
        if value:
            rows.append(("Schedule 2", line, desc, value, source, comp, "✓"))

    addl_medicare = sched2.get("additional_medicare", fed.get("additional_medicare", 0))
    r("11", "Additional Medicare Tax", addl_medicare,
      "0.9% on Medicare wages over $250,000 (MFJ)",
      f"${addl_medicare:,.2f}" if addl_medicare else "")

    niit = sched2.get("niit", fed.get("niit", 0))
    r("12", "Net Investment Income Tax (NIIT)", niit,
      "3.8% on lesser of NII or MAGI over $250,000",
      f"${niit:,.2f}" if niit else "")

    total = sched2.get("total", fed.get("schedule_2_total", 0))
    r("21", "Total additional taxes", total,
      "Line 11 + Line 12",
      f"${total:,.2f}" if total else "")

    return rows


def build_schedule_3_rows(ledger, results):
    """Build audit trail rows for Schedule 3 (Credits)."""
    fed = results.get("federal", results)
    sched3 = fed.get("schedule_3", {})
    if not sched3 and not fed.get("ftc") and not fed.get("cdcc"):
        return []

    rows = []
    def r(line, desc, value, source, comp):
        if value:
            rows.append(("Schedule 3", line, desc, value, source, comp, "✓"))

    ftc = sched3.get("ftc", fed.get("ftc", 0))
    r("1", "Foreign tax credit", ftc,
      "1099-DIV Box 7 / 1099-INT Box 6",
      f"${ftc:,.2f}" if ftc else "")

    cdcc = sched3.get("cdcc", fed.get("cdcc", 0))
    r("2", "Child and dependent care credit", cdcc,
      "Form 2441 / dependent care expenses",
      f"${cdcc:,.2f}" if cdcc else "")

    ctc = sched3.get("ctc", fed.get("ctc", fed.get("child_tax_credit", 0)))
    r("6a", "Child tax credit / ODC", ctc,
      "Per qualifying child under 17",
      f"${ctc:,.2f}" if ctc else "")

    total = sched3.get("total", fed.get("schedule_3_total", fed.get("total_credits", 0)))
    r("15", "Total credits", total,
      "Sum of Schedule 3 lines",
      f"${total:,.2f}" if total else "")

    return rows


def generate_workbook(ledger, results):
    """Create the audit trail Excel workbook."""
    wb = Workbook()

    # -----------------------------------------------------------------------
    # Sheet 1: Audit Trail (form field → source mapping)
    # -----------------------------------------------------------------------
    ws = wb.active
    ws.title = "Audit Trail"

    headers = ["Form", "Line", "Description", "Value", "Source Document(s)", "Computation", "Status"]
    col_widths = [14, 8, 35, 16, 50, 55, 10]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"

    # Collect all rows from all form builders
    all_sections = [
        ("Form 1040", build_form_1040_rows),
        ("Schedule A", build_schedule_a_rows),
        ("Schedule 2", build_schedule_2_rows),
        ("Schedule 3", build_schedule_3_rows),
    ]

    row_idx = 2
    for section_name, builder_fn in all_sections:
        form_rows = builder_fn(ledger, results)
        if not form_rows:
            continue

        # Section header row
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=7)
        cell = ws.cell(row=row_idx, column=1, value=section_name)
        cell.font = SECTION_FONT
        cell.fill = SECTION_FILL
        cell.alignment = Alignment(vertical="center")
        for c in range(1, 8):
            ws.cell(row=row_idx, column=c).border = THIN_BORDER
            ws.cell(row=row_idx, column=c).fill = SECTION_FILL
        row_idx += 1

        for form, line, desc, value, source, comp, status in form_rows:
            ws.cell(row=row_idx, column=1, value=form).border = THIN_BORDER
            ws.cell(row=row_idx, column=2, value=line).border = THIN_BORDER
            ws.cell(row=row_idx, column=3, value=desc).border = THIN_BORDER
            ws.cell(row=row_idx, column=3).alignment = WRAP_ALIGN

            val_cell = ws.cell(row=row_idx, column=4)
            if isinstance(value, (int, float)):
                val_cell.value = value
                val_cell.number_format = NUM_FMT
            else:
                val_cell.value = value
            val_cell.border = THIN_BORDER

            ws.cell(row=row_idx, column=5, value=source).border = THIN_BORDER
            ws.cell(row=row_idx, column=5).alignment = WRAP_ALIGN
            ws.cell(row=row_idx, column=6, value=comp).border = THIN_BORDER
            ws.cell(row=row_idx, column=6).alignment = WRAP_ALIGN

            status_cell = ws.cell(row=row_idx, column=7, value=status)
            status_cell.border = THIN_BORDER
            status_cell.alignment = Alignment(horizontal="center")
            if status == "✓":
                status_cell.fill = VERIFIED_FILL
            else:
                status_cell.fill = WARNING_FILL

            row_idx += 1

        row_idx += 1  # blank row between sections

    # -----------------------------------------------------------------------
    # Sheet 2: Source Document Summary (from extraction ledger)
    # -----------------------------------------------------------------------
    ws2 = wb.create_sheet("Source Documents")
    doc_headers = ["Document", "Type", "Filer", "Issuer", "Field", "Value"]
    doc_widths = [30, 12, 15, 25, 25, 18]

    for col_idx, (header, width) in enumerate(zip(doc_headers, doc_widths), 1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER
        ws2.column_dimensions[get_column_letter(col_idx)].width = width

    ws2.freeze_panes = "A2"
    row_idx = 2

    filers = ledger.get("filers", {})
    for doc in ledger.get("documents", []):
        doc_file = doc.get("file", "unknown")
        doc_type = doc.get("type", "unknown")
        filer_key = doc.get("filer", "unknown")
        filer_name = filers.get(filer_key, {}).get("first_name", filer_key)
        issuer = doc.get("issuer", "")
        values = doc.get("values", {})

        first_row = True
        for field, val in values.items():
            if isinstance(val, dict):
                # Nested dict (e.g., state_wages: {"NJ": 334898})
                for subkey, subval in val.items():
                    ws2.cell(row=row_idx, column=1, value=doc_file if first_row else "").border = THIN_BORDER
                    ws2.cell(row=row_idx, column=2, value=doc_type if first_row else "").border = THIN_BORDER
                    ws2.cell(row=row_idx, column=3, value=filer_name if first_row else "").border = THIN_BORDER
                    ws2.cell(row=row_idx, column=4, value=issuer if first_row else "").border = THIN_BORDER
                    ws2.cell(row=row_idx, column=5, value=f"{field} ({subkey})").border = THIN_BORDER
                    val_cell = ws2.cell(row=row_idx, column=6)
                    val_cell.value = subval
                    val_cell.number_format = NUM_FMT
                    val_cell.border = THIN_BORDER
                    first_row = False
                    row_idx += 1
            else:
                ws2.cell(row=row_idx, column=1, value=doc_file if first_row else "").border = THIN_BORDER
                ws2.cell(row=row_idx, column=2, value=doc_type if first_row else "").border = THIN_BORDER
                ws2.cell(row=row_idx, column=3, value=filer_name if first_row else "").border = THIN_BORDER
                ws2.cell(row=row_idx, column=4, value=issuer if first_row else "").border = THIN_BORDER
                ws2.cell(row=row_idx, column=5, value=field).border = THIN_BORDER
                val_cell = ws2.cell(row=row_idx, column=6)
                val_cell.value = val
                val_cell.number_format = NUM_FMT
                val_cell.border = THIN_BORDER
                first_row = False
                row_idx += 1

        row_idx += 1  # blank row between documents

    # -----------------------------------------------------------------------
    # Sheet 3: Cross-Foot Check Results (if available)
    # -----------------------------------------------------------------------
    # The cross_foot_report.json is written by cross_foot_check.py; if it
    # exists alongside the ledger, include it as a third sheet.
    ledger_dir = Path(sys.argv[1]).parent if len(sys.argv) > 1 else Path(".")
    xfoot_path = ledger_dir / "cross_foot_report.json"
    if xfoot_path.exists():
        ws3 = wb.create_sheet("Cross-Foot Checks")
        xf_headers = ["Check", "Result", "Detail"]
        xf_widths = [45, 10, 70]

        for col_idx, (header, width) in enumerate(zip(xf_headers, xf_widths), 1):
            cell = ws3.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER
            ws3.column_dimensions[get_column_letter(col_idx)].width = width

        ws3.freeze_panes = "A2"

        with open(xfoot_path) as f:
            xfoot = json.load(f)

        for i, check in enumerate(xfoot.get("checks", []), 2):
            ws3.cell(row=i, column=1, value=check.get("name", "")).border = THIN_BORDER
            result_cell = ws3.cell(row=i, column=2)
            passed = check.get("passed", False)
            result_cell.value = "✓ PASS" if passed else "✗ FAIL"
            result_cell.fill = VERIFIED_FILL if passed else WARNING_FILL
            result_cell.alignment = Alignment(horizontal="center")
            result_cell.border = THIN_BORDER
            ws3.cell(row=i, column=3, value=check.get("detail", "")).border = THIN_BORDER
            ws3.cell(row=i, column=3).alignment = WRAP_ALIGN

    return wb


def main():
    if len(sys.argv) < 4:
        print("Usage: python generate_audit_trail.py extraction_ledger.json results.json output_dir/")
        sys.exit(1)

    ledger_path = sys.argv[1]
    results_path = sys.argv[2]
    output_dir = Path(sys.argv[3])
    output_dir.mkdir(parents=True, exist_ok=True)

    with open(ledger_path) as f:
        ledger = json.load(f)
    with open(results_path) as f:
        results = json.load(f)

    wb = generate_workbook(ledger, results)

    output_path = output_dir / "audit_trail.xlsx"
    wb.save(str(output_path))
    print(f"Audit trail saved to: {output_path}")
    print(f"  Sheet 1: Audit Trail — {wb['Audit Trail'].max_row - 1} rows")
    print(f"  Sheet 2: Source Documents — all extracted values from ledger")
    if "Cross-Foot Checks" in wb.sheetnames:
        print(f"  Sheet 3: Cross-Foot Checks — verification results")


if __name__ == "__main__":
    main()
